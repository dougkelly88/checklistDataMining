import os, time, string, datetime
from openpyxl import Workbook, load_workbook

class Base(object):
    """ Base class for common properties"""
    def __init__(self):
        self.taskNumber = ""
        self.taskLabel = ""
        self.doneBy = ""
        self.startedAt = ""
        self.timeTaken = ""

    def basePopulate(self, ws, r):
        self.taskNumber = ws.cell(row = r, column = 2).value
        self.taskLabel = ws.cell(row = r, column =4).value
        self.doneBy = ws.cell(row = r, column = 5).value
        self.startedAt = ws.cell(row = r, column = 7).value
        self.timeTaken = ws.cell(row = r, column = 9).value

class PrintSampleData:

    def __init__(self, sampleName):

        self.sampleName = sampleName
        #TODO: check SOP version against script version - warning if mismatch
        self.generalTasks = GeneralTasks()
        self.electrodePrep = ElectrodePrep()
        self.printingSetup = PrintingSetup()
        self.printing = Printing()
        self.postPrinting = PostPrint()
        self.postIncubation = PostIncubation()

    def parseSampleID(self, sampleID):
        """ Outputs rig #, date """
        splitString = string.split(sampleID, "-")
        print(splitString)
        output = [""]*2
        if len(splitString)==3:
            output[0] = splitString[0][1]
            d=splitString[1][4:]
            m=splitString[1][2:4]
            y=splitString[1][0:2]
            output[1] = "%s/%s/20%s" % (d, m, y)
        else:
            #TODO: log warning about unexpected sample ID format, taking best guess
            output[0] = sampleID[1]
            y = sampleID[2:4]
            m = sampleID[4:6]
            d = sampleID[6:8]
            output[1] = "%s/%s/20%s" % (d, m, y)
        return output

class GeneralTasks:

    def __init__(self):

        self.roomTemperature = RoomTemperature()
        self.roomHumidity = RoomHumidity()
        self.tip = Tip()
        self.slide = Slide()
        self.oil = Oil()
        self.mix = Mix()

    def populate(self, ws):
        self.roomTemperature.populate(ws)
        self.roomHumidity.populate(ws)
        self.tip.populate(ws)
        self.slide.populate(ws)
        self.oil.populate(ws)
        self.mix.populate(ws)

class RoomTemperature(Base):

    def __init__(self):

        super(RoomTemperature, self).__init__()
        self.roomTemperature = ""

    def populate(self, ws):
        self.basePopulate(ws, 15)
        self.roomTemperature = ws['L15'].value

class RoomHumidity(Base):

    def __init__(self):

        super(RoomHumidity, self).__init__()
        self.roomHumidity = ""

    def populate(self, ws):
        self.basePopulate(ws, 17)
        self.roomHumidity = ws['L17'].value

class Tip(Base):

    def __init__(self):

        super(Tip, self).__init__()
        self.size = ""
        self.tipBatch = ""
        self.tipID = ""

    def populate(self, ws):
        self.basePopulate(ws, 19)
        self.size = ''.join([i for i in ws['L19'].value if i.isdigit()])
        self.tipBatch = ws['O19'].value
        self.tipID = ws['S19'].value

class Slide(Base):

    def __init__(self):

        super(Slide, self).__init__()
        self.CA = ""
        self.batch = ""
        self.slideID = ""

    def populate(self, ws):
        self.basePopulate(ws, 21)
        self.CA = ws['L21'].value
        self.batch = ws['O21'].value
        self.slideID = ws['S21'].value

class Oil(Base):

    def __init__(self):

        super(Oil, self).__init__()
        self.oilID = ""

    def populate(self, ws):
        self.basePopulate(ws, 23)
        self.oilID = ws['L23'].value

class Mix(Base):

    def __init__(self):

        super(Mix, self).__init__()
        self.type = ""
        self.claire633 = ""
        self.claire594 = ""
        self.claire532 = ""
        self.claire488 = ""

    def populate(self, ws):
        self.basePopulate(ws, 25)
        self.type = ws['L25'].value
        self.claire633 = ws['N25'].value
        self.claire594 = ws['P25'].value
        self.claire532 = ws['R25'].value
        self.claire488 = ws['T25'].value
        
class ElectrodePrep():

    def __init__(self):

        self.sonication = Sonication()
        self.washing = Washing()
        self.drying = Drying()
        self.flaming = Flaming()

    def populate(self, ws):
        self.sonication.populate(ws)
        self.washing.populate(ws)
        self.drying.populate(ws)
        self.flaming.populate(ws)

class Sonication(Base):

    def __init__(self):

        super(Sonication, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 27)
        

class Washing(Base):

    def __init__(self):

        super(Washing, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 29)

class Drying(Base):

    def __init__(self):

        super(Drying, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 31)

class Flaming(Base):

    def __init__(self):

        super(Flaming, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 33)

class PrintingSetup():

    def __init__(self):

        self.filling = Filling()
        self.mounting = Mounting()
        self.positionSlide = PositionSlide()
        self.positionOil = PositionOil()

    def populate(self, ws):
        self.filling.populate(ws)
        self.mounting.populate(ws)
        self.positionSlide.populate(ws)
        self.positionOil.populate(ws)

class Filling(Base):

    def __init__(self):

        super(Filling, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 35)

class Mounting(Base):

    def __init__(self):

        super(Mounting, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 37)

class PositionSlide(Base):

    def __init__(self):

        super(PositionSlide, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 39)

class PositionOil(Base):

    def __init__(self):

        super(PositionOil, self).__init__()
        self.lowHumidity = ""

    def populate(self, ws):
        self.basePopulate(ws, 41)
        self.lowHumidity = ws['L41'].value

class Printing():
    
    def __init__(self):

        self.moveIntoOil = MoveIntoOil()
        self.humidifierHigh = HumidifierHigh()
        self.optimisePrinting = OptimisePrinting()
        self.pPrint = PPrint()

    def populate(self, ws):
        self.moveIntoOil.populate(ws)
        self.humidifierHigh.populate(ws)
        self.optimisePrinting.populate(ws)
        self.pPrint.populate(ws)

class MoveIntoOil(Base):

    def __init__(self):

        super(MoveIntoOil, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 43)

class HumidifierHigh(Base):

    def __init__(self):

        super(HumidifierHigh, self).__init__()
        self.highHumidity = ""

    def populate(self, ws):
        self.basePopulate(ws, 45)
        self.highHumidity = ws['L45'].value

class OptimisePrinting(Base):

    def __init__(self):

        super(OptimisePrinting, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 47)

class PPrint(Base):

    def __init__(self):

        super(PPrint, self).__init__()
        self.dwell = ""
        self.step = ""
        self.voltage = ""
        self.frequency = ""
        self.pressure = ""
        

    def populate(self, ws):
        self.basePopulate(ws, 49)
        self.dwell = ws['L50'].value
        self.step = ws['N50'].value
        self.voltage = ws['P50'].value
        self.frequency = ws['R50'].value
        self.pressure = ws['T50'].value

class PostPrint():
    def __init__(self):

        self.transferToOven = TransferToOven()
        self.recoverMix = RecoverMix()
        self.tipSize = TipSize()

    def populate(self, ws):
        self.transferToOven.populate(ws)
        self.recoverMix.populate(ws)
        self.tipSize.populate(ws)

class TransferToOven(Base):
    def __init__(self):

        super(TransferToOven, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 52)

class RecoverMix(Base):
    def __init__(self):

        super(RecoverMix, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 54)

class TipSize(Base):
    def __init__(self):

        super(TipSize, self).__init__()
        tipSizeUm = ""

    def populate(self, ws):
        self.basePopulate(ws, 56)
        self.tipSizeUm = ws['L56'].value

class PostIncubation():
    def __init__(self):

        self.coolDown = CoolDown()
        self.checkWater = CheckWater()
        self.waitTime = WaitTime()
        self.measurePushThrough = MeasurePushThrough()

    def populate(self, ws):
        self.coolDown.populate(ws)
        self.checkWater.populate(ws)
        self.waitTime.populate(ws)
        self.measurePushThrough.populate(ws)

class CoolDown(Base):
    def __init__(self):

        super(CoolDown, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 58)

class CheckWater(Base):
    def __init__(self):

        super(CheckWater, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 60)

class WaitTime(Base):
    def __init__(self):

        super(WaitTime, self).__init__()

    def populate(self, ws):
        self.basePopulate(ws, 62)

class MeasurePushThrough(Base):
    def __init__(self):

        super(MeasurePushThrough, self).__init__()
        self.claire633 = ""
        self.claire594 = ""
        self.claire532 = ""
        self.claire488 = ""
        
    def populate(self, ws):
        self.basePopulate(ws, 64)
        self.claire633 = ws['L64'].value
        self.claire594 = ws['N64'].value
        self.claire532 = ws['P64'].value
        self.claire488 = ws['R64'].value

def outputData(psd, gspd):

    output=[]
    # Perform parsing and caclulating tasks
    parsedSampleID = psd.parseSampleID(psd.sampleName)
    output.append(parsedSampleID[1])                            # date
    output.append("")                                           # experiment ID
    output.append(psd.sampleName)                               # sample name
    output.append("")                                           # protocol version
    output.append("Rig%s" % parsedSampleID[0])                  # Rig
    output.append(psd.printing.pPrint.doneBy)                   # Printer
    output.append(aMPM(psd.printingSetup.filling.startedAt))    # AM/PM
    output.append("")                                           # Run score
    output.append("")                                           # Basecalling?
    output.append(psd.generalTasks.slide.CA)                    # Slide CA
    output.append(psd.generalTasks.slide.batch)                 # Slide batch
    output.append(psd.generalTasks.tip.size)                    # Tip size
    output.append(psd.generalTasks.tip.tipBatch)                # Tip batch
    output.append("")                                           # blank column
    #output.append(gpsd.oilWaterRest)                            # Oil/water rest time
    output.append("")
    #output.append(gpsd.oilWaterVortex)                          # Oil/water vortex time
    output.append("")
    output.append("")                                           # Oil/surfactant prepared by weight?
    #output.append(gpsd.surfactantConcn)                         # Surfactant concentration
    output.append("")
    output.append(psd.generalTasks.roomTemperature.roomTemperature) # Room temp
    output.append(psd.generalTasks.roomHumidity.roomHumidity)   # Room humidity
    output.append(psd.printingSetup.positionOil.lowHumidity)    # Low humidity
    output.append(psd.printing.humidifierHigh.highHumidity)     # High humidity
    output.append("")                                           # empty
    output.append(psd.printing.pPrint.timeTaken)                # Print time
    output.append("")                                           # Array quality
    output.append("")                                           # Droplet density
    output.append(psd.printing.pPrint.voltage)                  # Print voltage
    output.append(psd.printing.pPrint.frequency)                # Print frequency
    output.append(psd.printing.pPrint.dwell)                    # Dwell time
    output.append(psd.printing.pPrint.step)                     # Step size
    output.append(psd.printing.pPrint.pressure)                 # Pressure
    output.append("")                                           # blank
    output.append("")                                           # Air con
    output.append("")                                           # Door
    output.append("")                                           # blank
    output.append(psd.generalTasks.mix.type)                    # Mix
    output.append(psd.generalTasks.mix.claire633)               # 633 bulk
    output.append(psd.generalTasks.mix.claire594)               # 594 bulk
    output.append(psd.generalTasks.mix.claire532)               # 532 bulk
    output.append(psd.generalTasks.mix.claire488)               # 488 bulk
    output.append(psd.postIncubation.measurePushThrough.claire633)  # 633 push through
    output.append(psd.postIncubation.measurePushThrough.claire594)  # 594 push through
    output.append(psd.postIncubation.measurePushThrough.claire532)  # 532 push through
    output.append(psd.postIncubation.measurePushThrough.claire488)  # 488 push through                      
    
    print(output)
    

def aMPM(startTime):
    print('AMPMcheck')
    print(startTime)
    if isinstance(startTime, basestring):
        print("startTime is a string")
        a = int(startTime[0:2])
        if a < 12:
            return "AM"
        else:
            return "PM"
    elif isinstance(startTime, datetime.time):
        if startTime.hour < 12:
            return "AM"
        else:
            return "PM"
    else:
        print("Type not handled!")
        print(type(startTime))

def outputDataToSpreadsheet(outputData):

    # identify analysis spreadsheet and whether it can be opened for writing:
    wb = load_workbook('Book1.xlsx')
    wb
    

if __name__ == "__main__":

    wb = load_workbook('Z:\\SOPs\\Completed Checklists\\Data\\Printing\\Printing 1 2015-10-09 1130.xlsm')
    ws = wb.active
    
    psd = PrintSampleData(ws['E8'].value)
    #gspd = GeneralSamplePrepData()
    gspd = None

    print(psd.sampleName)
    psd.generalTasks.populate(ws)
    psd.electrodePrep.populate(ws)
    psd.printingSetup.populate(ws)
    psd.printing.populate(ws)
    psd.postPrinting.populate(ws)
    psd.postIncubation.populate(ws)


    # Perform parsing and caclulating tasks
    

    #testing
    outputData(psd, gspd)
    #print(psd.generalTasks.roomTemperature.taskNumber)
    #print(psd.generalTasks.roomTemperature.taskLabel)
    #output = psd.parseSampleID(psd.sampleName)
    #print(output)
    #output = psd.parseSampleID("P1150607A")
    #print(output)
    #print(aMPM("11:30"))
    #print('BREAK')
    
##    print(psd.generalTasks.roomTemperature.roomTemperature)
##    print(psd.generalTasks.roomTemperature.doneBy)
##    print(psd.generalTasks.mix.doneBy)
##    print(psd.printingSetup.positionOil.doneBy)
##    print(psd.printingSetup.positionOil.startedAt)
##    print(psd.printingSetup.positionOil.lowHumidity)
##
##    print('BREAK')
##    print(psd.printing.pPrint.frequency)
##    print(psd.postIncubation.measurePushThrough.claire633)
##    print(psd.postPrinting.recoverMix.doneBy)
##    print(psd.postPrinting.recoverMix.startedAt)
