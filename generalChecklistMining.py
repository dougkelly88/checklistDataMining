﻿import os, time, string, datetime, tkFileDialog, ast, webbrowser, sys, traceback, re, numbers
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from Tkinter import *
from tkMessageBox import *
import tkSimpleDialog
import requests, gspread
from oauth2client.client import SignedJwtAssertionCredentials

# Constants
MODE_APPEND = 0
MODE_NEW = 1
MODE_UPDATEWEBFROMCL = 2
CHECKLIST_FILE_FORMAT = '.xlsm'

# Error messages
NOT_YET_SUPPORTED = "Sorry, this feature is not yet supported!"
INVALID_PATH = "No (valid) path selected!"
FILES_LENGTH_ZERO = "No (valid) files found in this location!"

class outputSelectionDialog(object):
    """ Dialog box allowing selection of output params"""
    def __init__(self, internalData):
        self.data = internalData
        root = self.root = Tk()
        self.all = IntVar()
        self.sizex = 400
        self.sizey = 600
        self.posx  = 100
        self.posy  = 100
        root.wm_geometry("%dx%d+%d+%d" % (self.sizex, self.sizey, self.posx, self.posy))
        root.title('Select data to export...')
        frm_1 = Frame(root)
        frm_1.pack(ipadx=2, ipady=2)
        self.canvas = Canvas(frm_1)
        frm_2 = Frame(self.canvas)
        scrlbar = Scrollbar(frm_1,orient="vertical",command=self.canvas.yview)
        scrlbar.pack(side="right",fill="y")
        self.canvas.configure(yscrollcommand=scrlbar.set)
        self.canvas.pack(side="left")
        self.canvas.create_window((0,0),window=frm_2,anchor='nw')
        frm_2.bind("<Configure>",self.myfunction)

        i = 0
        self.vars = []
        chkAll = Checkbutton(frm_2, text = 'All', variable = self.all)
        chkAll.pack(padx = 0, pady = 5, anchor = W)
        chkAll['command'] = self.c2_action
        if isinstance(self.data, Printing):
            for t in self.data.tasks:
                txt = "%s: %s" % (t.taskCategory, t.taskLabel)
                var = IntVar()
                chk = Checkbutton(frm_2, text = txt, variable = var)
                i = i+1
                chk.pack(padx = 0, pady = 5, anchor = W)
                chk['command'] = self.c1_action
                self.vars.append(var)
        else:
            for t in self.data.tasks:
                txt = "%s" % (t.taskLabel)
                var = IntVar()
                chk = Checkbutton(frm_2, text = txt, variable = var)
                i = i+1
                chk.pack(padx = 0, pady = 5, anchor = W)
                chk['command'] = self.c1_action
                self.vars.append(var)
        btn_1 = Button(frm_2, width=8, text='OK')
        btn_1['command'] = self.b1_action
        btn_1.pack(anchor = CENTER)
        
    def myfunction(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"),width=self.sizex,height=self.sizey)
    
    def b1_action(self, event=None):
        #print('quitting...')
        self.root.quit()

    def c2_action(self, event = None):
        i = 0
        for v in self.vars:
            if (self.all.get() == 1):
                self.data.tasks[i].output = True
                v.set(1)
                #print(self.data.tasks[i].taskLabel)
                #print(self.data.tasks[i].output)
            else:
                self.data.tasks[i].output = False
                v.set(0)
            print(self.data.tasks[i].taskLabel)
            print(self.data.tasks[i].output)
            i = i+1


    def c1_action(self, event = None):
        i = 0
        for v in self.vars:
            if (v.get() == 1):
                self.data.tasks[i].output = True
            else:
                self.data.tasks[i].output = False
            print(self.data.tasks[i].taskLabel)
            print(self.data.tasks[i].output)
            i = i+1

class ChecklistBase(object):
    """ Base class for checklist-level common properties"""
    def __init__(self):
        self.sOP = ""
        self.sOPVersion = ""
        self.date = ""
        self.qCPerson = ""
        self.tasks = []
        self.path = ""
        self.output = False

    def returnTaskByLabel(self, label):
        for task in self.tasks:
            if (task.taskLabel.lower() == label.lower()):
                return task

    def populateTasks(c, ws):
        print('Populating...')
        category = ""
        for row in ws.iter_rows('B1:B100'):
            taskLabel = None
            for cell in row:
                if isinstance(c, Printing):
                    if (isinstance(cell.value, long)):
                        r = row[0].row
                        taskNumber = cell.value    
                        print("task number")
                        print(taskNumber)
                        if (ws.cell(row = r, column = 3).value != None):
                            category = ws.cell(row = r, column = 3).value
                        taskLabel = ws.cell(row = r, column = 4).value
                        print(taskLabel)
                else:
                    if (isinstance(cell.value, long)) | (isinstance(cell.value, float)):
                        r = row[0].row
                        taskNumber = cell.value
                        print(taskNumber)
                        if (ws.cell(row = r, column = 3).value != None):
                            taskLabel = ws.cell(row = r, column = 3).value
                if (taskLabel != None):
                    print(taskLabel)
                    t = c.identifyCorrectClass(taskLabel)
                    t.taskLabel = taskLabel
                    t.taskCategory = category
                    t.taskNumber  = taskNumber
                    t.populate(ws, r)
                    print(t)
                    #print(vars(t))
                    c.tasks.append(t)

class TaskBase(object):
    """ Base class for task-level common properties"""
    def __init__(self):
        self.taskNumber = ""
        self.taskLabel = ""
        self.taskCategory = ""
        self.doneBy = ""
        self.startedAt = ""
        self.timeTaken = ""    
        self.output = False       
        self.notes = []

    def populate(self, ws, r):
        self.doneBy = ws.cell(row = r, column = 5).value
        self.startedAt = ws.cell(row = r, column = 7).value
        self.timeTaken = ws.cell(row = r, column = 9).value
        self.output = False

    def returnTaskByLabel(self, label):
        if (taskLabel == label):
            return self

class Printing(ChecklistBase):
    """Class containing all variables and methods relating to the printing checklist"""
    def __init__(self, path):
        # Initialise general checklist class
        super(Printing, self).__init__()
        self.path = path
        # Initialise specialised print checklist variables
        self.sampleName = ""
        self.experimenter = ""
        
        # Initialise specialised print checklist variables that are not explicitly in checklist but may be derived from checklist entries
        self.printDate = ""
        self.printRig = ""

    class PrintTask(TaskBase):
        def __init__(self):
            self.dwell = ""
            self.step = ""
            self.voltage = ""
            self.freq = ""
            self.pressure = ""
            self.dcOffset = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for rowind in range(2):
                for col in range(26):
                    c = ws.cell(row = r+rowind, column = col).value
                    if (isinstance(c, basestring)):
                        if ("dwell" in c.lower()):
                            self.dwell = ws.cell(row = r+rowind, column = col+1).value
                        if ("step" in c.lower()):
                            self.step = ws.cell(row = r+rowind, column = col+1).value
                        if ("voltage" in c.lower()):
                            self.voltage = ws.cell(row = r+rowind, column = col+1).value
                        if ("freq" in c.lower()):
                            self.freq = ws.cell(row = r+rowind, column = col+1).value
                        if ("pressure" in c.lower()):
                            self.pressure = ws.cell(row = r+rowind, column = col+1).value
                        if ("offset" in c.lower()):
                            self.dcOffset = ws.cell(row = r+rowind, column = col+1).value

    class FillingTipTask(TaskBase):
        def __init__(self):
            self.mixVolume = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("mix volume" in c.lower()):
                        self.mixVolume = ws.cell(row = r, column = col+1).value
                        print("fill volume = %s" % self.mixVolume)

    class MixTask(TaskBase):
        def __init__(self):
            self.id = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("id" in c.lower()):
                        self.id = ws.cell(row = r, column = col+1).value
                    
    class OvenTask(TaskBase):
        def __init__(self):
            self.type = ""
            self.temperature = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("oven type" in c.lower()):
                        self.type = ws.cell(row = r, column = col+1).value
                    if ("temp" in c.lower()):
                        s = ws.cell(row = r, column = col+1).value
                        if (isinstance(s, basestring)):
                            f = s.split("C", 1)
                        else:
                            f = s
                        self.temperature = f

    class OilTask(TaskBase):
        def __intit__(self):
            self.aliquote = ""
            self.id = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("ID" in c):
                        self.id = ws.cell(row = r, column = col+1).value
                    if ("Aliquote" in c):
                        self.aliquote = ws.cell(row = r, column = col+1).value

    class BoxTask(TaskBase):
        def __init__(self):
            self.bottom = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26): 
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("Bottom" in c):
                        self.bottom = ws.cell(row = r, column = col+1).value
                        print("bottom=")
                        print(self.bottom)

    class BulksTask(TaskBase):
        def __init__(self):
            self.type = ""
            self.claire700 = ""
            self.claire655 = ""
            self.claire594 = ""
            self.claire532 = ""
            self.claire488 = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)
            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("633nm" in c) or ("655nm" in c):
                        self.claire655 = ws.cell(row = r, column = col+1).value
                    if ("700nm" in c):
                        self.claire700 = ws.cell(row = r, column = col+1).value
                    if ("594nm" in c):
                        self.claire594 = ws.cell(row = r, column = col+1).value
                    if ("532nm" in c):
                        self.claire532 = ws.cell(row = r, column = col+1).value
                    if ("488nm" in c):
                        self.claire488 = ws.cell(row = r, column = col+1).value
                    if ("type" in c.lower()):
                        self.type = ws.cell(row = r, column = col+1).value

    class TipTask(TaskBase):
        def __init__(self):
            self.size = ""
            self.batch = ""
            self.ID = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("Size" in c):
                        self.size = ws.cell(row = r, column = col+1).value
                    if ("batch" in c.lower()):
                        if isinstance(ws.cell(row = r, column = col+3).value, numbers.Number):
                            self.batch = '%s-%02d' % (ws.cell(row = r, column = col+1).value, ws.cell(row = r, column = col+3).value)
                        else:
                            self.batch = '%s-NONE' % (ws.cell(row = r, column = col+1).value)

                    if ("ID" in c):
                        self.ID = ws.cell(row = r, column = col+1).value

    class SlideTask(TaskBase):
        def __init__(self):
            self.CA = ""
            self.batch = ""
            self.ID = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    print(c)
                    if ("CA" in c):
                        self.CA = ws.cell(row = r, column = col+1).value
                        print("slide CA")
                        print(self.CA)
                    if ("batch" in c.lower()):
                        self.batch = ws.cell(row = r, column = col+1).value
                        print("slide batch")
                        print(self.batch)
                    if ("#" in c):
                        self.ID = ws.cell(row = r, column = col+1).value
                        print("slide number")
                        print(self.ID)

    class HumidityTask(TaskBase):
        def __init__(self):
            self.humidity = ""
            self.oilVolume = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("Volume" in c):
                        self.oilVolume = ws.cell(row = r, column = col+1).value
                    if ("Humidity" in c):
                        self.humidity = ws.cell(row = r, column = col+1).value

    class TemperatureTask(TaskBase):
        def __init__(self):
            self.temperature = ""

        def populate(self, ws, r):
            TaskBase.populate(self, ws, r)
            #print(self.output)
            self.temperature = ws.cell(row = r, column = 12).value
            self.notes = ws.cell(row = r, column = 14).value
            #print("Temperature task")
            #print(vars(self))

    def populatePrintingClass(self, ws):
        """Populate checklist-level data for the Printing case"""
        self.sOP = ws['E2'].value
        self.sOPVersion = ws['E4'].value
        self.date = ws['E6'].value
        self.sampleName = ws['E8'].value
        self.experimenter = ws['E10'].value
        self.qCPerson = ws['E12'].value
        self.printDate = self.parseSampleID(self.sampleName)[1]
        self.printRig = "P%s" % self.parseSampleID(self.sampleName)[0]

    def parseSampleID(self, sampleID):
        """ Outputs rig #, date """
        output = [""]*2
        if isinstance(sampleID, basestring):
            dashes = sampleID.count('-')
            if (dashes == 2):
                splitString = string.split(sampleID, "-")
            #if len(splitString)==3:
                output[0] = splitString[0][1]
                d=splitString[1][4:]
                m=splitString[1][2:4]
                y=splitString[1][0:2]
                
            elif (dashes == 1):
                output[0] = sampleID[1]
                y = sampleID[3:5]
                m = sampleID[5:7]
                d = sampleID[7:9]

            else:
                #TODO: log warning about unexpected sample ID format, taking best guess
                output[0] = sampleID[1]
                y = sampleID[2:4]
                m = sampleID[4:6]
                d = sampleID[6:8]
                
            output[1] = "%s%s%s" % (y, m, d)
        return output

    def identifyCorrectClass(self, description):
        description = description.lower()
        print("description=")
        print(description)
        if ("humidity" in description) or ("humidifier" in description):
            return Printing.HumidityTask()
        if ("temperature" in description):
            return Printing.TemperatureTask()
        if (description == "slide (note batch and id)"):
            return Printing.SlideTask()
        if (description == "tip (note size and #)") or ("tip size" in description):
            return Printing.TipTask()
        #if (description == "mix (note id)") or ("push-through" in description):
        #    return Printing.BulksTask()
        if (description == "print"):
            return Printing.PrintTask()
        if (description == "oil (note batch and id)"):
            return Printing.OilTask()
        if ("box check intact" in description):
            print("found box check intact")
            return Printing.BoxTask()
        if ("oven" in description):
            return Printing.OvenTask()
        if (description == "fill tip"):
            return Printing.FillingTipTask()
        if (description == "mix (note id)"):
            return Printing.MixTask()
        else:
            return TaskBase()

class PrintingPrep(ChecklistBase):
    """Class containing all variables and methods relating to the printing prep checklist"""
    def __init__(self, path):
        # Initialise general checklist class
        super(PrintingPrep, self).__init__()
        self.path = path

        # Initialise specialised printing prep vars
        self.oilIDs = []

    def populatePrintingPrepClass(self, ws):
        """Populate checklist-level data for the Printing case"""
        self.sOP = ws['D2'].value
        self.sOPVersion = ws['D4'].value
        self.date = ws['D6'].value
        self.qCPerson = ws['D8'].value
        
    class stockABILParaffinTask(TaskBase):
        def __init__(self):
            self.surfactantConcn = ""

        def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)
            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("surfactant" in c.lower()):
                        self.surfactantConcn = ws.cell(row = r, column = col+3).value
                        print(self.surfactantConcn)

    class oilWaterMixTask(TaskBase):
        def __init__(self):
            self.date = ""
            self.time = ""

        def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("date" in c.lower()):
                        self.date = ws.cell(row = r, column = col+1).value
                    if ("time" in c.lower()):
                        self.time = ws.cell(row = r, column = col+1).value
    
    # Condense hydrate, settle and rotate, all of which have same vars, into a single class?
    class hydrateOilTask(TaskBase):
        def __init__(self):
            self.duration = ""
            self.batch = ""

        def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)

            self.batch = returnBatchNumber(self.taskLabel)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("duration" in c.lower()):
                        self.duration = ws.cell(row = r, column = col+1).value
     
    class settleOilTask(TaskBase):
         def __init__(self):
            self.duration = ""
            self.batch = ""

         def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)

            self.batch = returnBatchNumber(self.taskLabel)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("duration" in c.lower()):
                        self.duration = ws.cell(row = r, column = col+1).value
                                       
    class addABILToHydratedOilTask(TaskBase):
        def __init__(self):
            self.surfactantConcn = ""
            self.batch = ""

        def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)
            
            self.batch = returnBatchNumber(self.taskLabel)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("surfactant" in c.lower()):
                        self.surfactantConcn = ws.cell(row = r, column = col+3).value
                        if isinstance(self.surfactantConcn, float):
                            self.surfactantConcn = self.surfactantConcn * 100;
                        print(self.surfactantConcn)

    class rotateTask(TaskBase):
         def __init__(self):
            self.duration = ""
            self.batch = ""

         def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)

            self.batch = returnBatchNumber(self.taskLabel)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("duration" in c.lower()):
                        self.duration = ws.cell(row = r, column = col+1).value

    class aliquoteTask(TaskBase):
        def __init__(self, outer):
            self.oilID = ""
            self.batch = ""
            self.outerInstance = outer

        def populate(self, ws, r):
            #TaskBase.populate(self, ws, r)

            self.batch = returnBatchNumber(self.taskLabel)

            for col in range(26):
                c = ws.cell(row = r, column = col).value
                if (isinstance(c, basestring)):
                    if ("id" in c.lower()):
                        self.oilID = ws.cell(row = r, column = col+1).value
                        self.outerInstance.oilIDs.append(self.oilID)

            if (self.batch == 1):
                self.outerInstance.date = self.oilID[0:6]

    def identifyCorrectClass(self, description):
        #print('IDing class...')
        #print(description)
        if (description == "Stock ABIL in Paraffin prep"):
            return PrintingPrep.stockABILParaffinTask()
        if (description == "Prep Oil/water mix"):
            return PrintingPrep.oilWaterMixTask()
        if ("hydrate" in description.lower()):
            return PrintingPrep.hydrateOilTask()
        if ("Settling time" in description):
            return PrintingPrep.settleOilTask()
        if ("Mix with 5% ABIL" in description):
            return PrintingPrep.addABILToHydratedOilTask()
        if ("rotate" in description.lower()):
            return PrintingPrep.rotateTask()
        if ("aliquote" in description.lower()):
            return PrintingPrep.aliquoteTask(self);
        else:
            return TaskBase()

def authenticate_google_docs():
    #f = file(os.path.join('C:/Users/d.kelly/Desktop/Python/bulkSummariser-f4e730f107d4.p12'), 'rb')
    f = file(os.path.join('//base4share/share/Doug/2015/checklistDataMining/bulkSummariser-f4e730f107d4.p12'), 'rb')
    SIGNED_KEY = f.read()
    f.close()
    scope = ['https://spreadsheets.google.com/feeds', 'https://docs.google.com/feeds', 'https://www.googleapis.com/auth/gmail.send']
    credentials = SignedJwtAssertionCredentials('d.kelly@base4.co.uk', SIGNED_KEY, scope)

    data = {
        'refresh_token' : '1/J9DflNyNnx2J9WnWtHsKnH9YDTTBTuoz3tJtnSXRtLc',
        'client_id' : '326340397307-lgicfbcjmu9863pjjkfrn0c5dviqsb42.apps.googleusercontent.com',
        'client_secret' : 'xeNtki_QWyoLaGEjYqtzF5si',
        'grant_type' : 'refresh_token',
    }

    r = requests.post('https://accounts.google.com/o/oauth2/token', data = data)
    credentials.access_token = ast.literal_eval(r.text)['access_token']

    gc = gspread.authorize(credentials)
    return gc

def chooseFolder(initialdir, prompt):
    try:
        import tkFileDialog
    except:
        from tkinter import filedialog as tkFileDialog

    root = Tk()
    try:
        options = {}
        options['initialdir']=initialdir
        file = tkFileDialog.askdirectory(**options)
        print(file)
        
        if file == "":
            errorHandler(INVALID_PATH)

    except IOError:
        errorHandler(INVALID_PATH)

    root.destroy()
    return file

def chooseOutputFile(initialdir, initialfile):
    try:
        root = Tk()
        options = {}
        options['initialdir']=initialdir
        options['initialfile'] = initialfile
        options['defaultextension'] = '.xls'
        options['filetypes'] = [('Excel spreadsheet', '.xlsx'), ('Comma separated variable file', '.csv')]
        options['title'] = 'Choose an output file to save the summary data to...'
        file = tkFileDialog.asksaveasfilename(**options)
        print(file)
        
        if file == "":
            errorHandler(INVALID_PATH)

    except IOError:
        errorHandler(INVALID_PATH)

    root.destroy()
    return file

def errorHandler(message):
    print(message)
    showerror("Error!", message)
    # TODO: log file? send email?
    exit()

def numberToLetters(q):
    q = q - 1
    result = ''
    while q >= 0:
        remain = q % 26
        result = chr(remain+65) + result;
        q = q//26 - 1
    return result
                
def formatToGS(p, gws):

    # cross reference with PrintingPrep sheet here!

    # find row, or make new row
    sampleNamesFromGS = gws.col_values(1)
    date = datetime.datetime.strptime(p.printDate, '%y%m%d')
    fmt = ('%d/%m/%Y')
    if p.sampleName in sampleNamesFromGS:
        row = sampleNamesFromGS.index(p.sampleName) + 1
        gws.update_cell(row, 2, date.strftime(fmt))
    else:
        row = len(sampleNamesFromGS) + 1
        gws.update_cell(row, 1, p.sampleName)
        gws.update_cell(row, 2, date.strftime(fmt))

    headings = gws.row_values(1)
    splitPath = os.path.split(p.path)
    filename = os.path.splitext(splitPath[1])
    fmt = ('%Y\%B\%d')
    datePath = date.strftime(fmt)
    pdfPath = os.path.join("\\\\base4share\share\SOPs\Completed Checklists", datePath, (filename[0]).replace("Printing 1", "Printing") +  '.pdf')
    gws.update_cell(row, headings.index("Completed Checklist Link") + 1, pdfPath)

    first_col = numberToLetters(headings.index("Protocol version") + 1)
    last_col = numberToLetters(headings.index("Mix number") + 1) 
    cellrange = '%s%d:%s%d' % (first_col, row, last_col, row)
    cells = gws.range(cellrange)
    headings = headings[headings.index("Protocol version"):(headings.index("Mix number") + 1)]

    for heading in headings:
        # should each of these be surrounded by its own try-catch for robustness?!
        if (heading == "Protocol version"):
            cells[headings.index(heading) ].value = 'v %1.2f' % p.sOPVersion
        if (heading == "Rig"):
            cells[headings.index(heading) ].value = p.printRig
        if (heading == "Printer"):
            cells[headings.index(heading) ].value = p.experimenter
        #if (heading == "Slide CA"):
        #    cells[headings.index(heading) ].value = p.returnTaskByLabel("Slide (note batch and ID)").CA
        if (heading == "Slide batch"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Slide (note batch and ID)").batch
        if (heading == "Tip size"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Tip (note size and #)").size
        if (heading == "Tip Batch"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Tip (note size and #)").batch
        if (heading == "Tip ID"):
            cells[headings.index(heading) ].value = "%s%s" % (p.printDate, p.returnTaskByLabel("Tip (note size and #)").ID)
        if (heading == "Room Temperature"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Note room temperature").temperature
        if (heading == "Room Humidity"):
            if (p.returnTaskByLabel("Note room humidity").humidity != None):
                cells[headings.index(heading) ].value = p.returnTaskByLabel("Note room humidity").humidity * 100
        #if (heading == "Humidity low"):
        #    if (p.returnTaskByLabel("Position Oil + Turn humidifier to low").humidity != None):
        #        cells[headings.index(heading) ].value = p.returnTaskByLabel("Position Oil + Turn humidifier to low").humidity * 100
        if (heading == "Humidity high"):
            try:
                if (p.returnTaskByLabel("Insert well, check box humidity >80%").humidity != None):
                    cells[headings.index(heading) ].value = p.returnTaskByLabel("Insert well, check box humidity >80%").humidity * 100
            except:
                if (p.returnTaskByLabel("Turn humidifier to high").humidity != None):
                    cells[headings.index(heading) ].value = p.returnTaskByLabel("Turn humidifier to high").humidity * 100
        if (heading == "Print time"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Print").timeTaken
        if (heading == "Print voltage (AC) / V x 100"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Print").voltage
        if (heading == "Voltage frequency (sine) /Hz"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Print").freq
        if (heading == "Dwell time"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Print").dwell
        if (heading == "Step size"):
            cells[headings.index(heading) ].value = p.returnTaskByLabel("Print").step
        if (heading == "Pressure [kPa]"):
            cells[headings.index(heading)].value = p.returnTaskByLabel("Print").pressure
        if (heading == "Rig Box open/closed"):
                cells[headings.index(heading)].value = p.returnTaskByLabel("Box check intact").bottom
        if (heading == "Tip fill/ul"):
            cells[headings.index(heading)].value = p.returnTaskByLabel("Fill tip").mixVolume
        if (heading == "oil volume/ul"):
            cells[headings.index(heading)].value = p.returnTaskByLabel("Insert well, check box humidity >80%").oilVolume
        if (heading == "Print voltage (DC) / V x 100"):
            cells[headings.index(heading)].value = p.returnTaskByLabel("Print").dcOffset
        if (heading == "incubation"):
            if (p.returnTaskByLabel("Transfer to oven").temperature is not None):
                str = ''
                print(p.returnTaskByLabel("Transfer to oven").temperature)
                if isinstance(p.returnTaskByLabel("Transfer to oven").temperature, numbers.Number):
                    #print('isnum')
                    t = float(p.returnTaskByLabel("Transfer to oven").temperature)
                    str = "%s @ %02.1f" % (p.returnTaskByLabel("Transfer to oven").type, t)
                elif isinstance(p.returnTaskByLabel("Transfer to oven").temperature, basestring):
                    #str = p.returnTaskByLabel("Transfer to oven").temperature.encode('latin1')
                    #print('isbasestr')
                    str = "%s @ %s" % (p.returnTaskByLabel("Transfer to oven").type, p.returnTaskByLabel("Transfer to oven").temperature.encode('latin1'))
                elif isinstance(p.returnTaskByLabel("Transfer to oven").temperature, list):
                     #print('islist')
                     str = "%s @ %s" % (p.returnTaskByLabel("Transfer to oven").type, p.returnTaskByLabel("Transfer to oven").temperature[0].encode('latin1'))
                print(type(p.returnTaskByLabel("Transfer to oven").temperature))
                print(str)
                cells[headings.index(heading)].value = str
        if (heading == "oil/surfactant batch ID"):
            if p.returnTaskByLabel("Oil (note batch and ID)").aliquote is not None:
                v = '%s-%d' % (p.returnTaskByLabel("Oil (note batch and ID)").id, p.returnTaskByLabel("Oil (note batch and ID)").aliquote)
            else:
                v = p.returnTaskByLabel("Oil (note batch and ID)").id
            cells[headings.index(heading)].value = v
        if (heading == "Mix number"):
            cells[headings.index(heading)].value = p.returnTaskByLabel("Mix (note ID)").id

    gws.update_cells(cells)
    #raw_input('press enter to continue...')
    webbrowser.open('https://docs.google.com/spreadsheets/d/1gbs1BvFy69ISbnAYRfxLj74tpYalQjAWPbhT2hsr_Z8/edit#gid=1149687153', 2, True)

def uploadToHiddenGS(p, gsh):

    # for debug, pass authentication and open dummy sheet here; for real thing, pass gsh
    #gsh = gc.open("Dummy sample register")
    gws = gsh.worksheet("Oil prep")
    headings = gws.row_values(1)
    
    first_col = numberToLetters(headings.index("Date") + 1)
    last_col = numberToLetters(headings.index("final oil/surfactant mix conc %w?") + 1) 
    
    # add all conceivable aliquot ID-aliquot number combinations
    # just in case, add possibility of no aliquot number being employed
    for aliquot_no in range(11):
        for oilID in p.oilIDs:
            if oilID is not None:
                idFromGS = gws.col_values(headings.index("ID") + 1)
                row = len(idFromGS) + 1
                print('row=%d' % row) 
                cellrange = '%s%d:%s%d' % (first_col, row, last_col, row)
                cells = gws.range(cellrange)

        
                batch = p.oilIDs.index(oilID) + 1

                if (aliquot_no == 10):
                    id = '%s' % (oilID)
                else:
                    id = '%s-%d' % (oilID, aliquot_no+1)
                cells[headings.index("ID")].value = id
                date = datetime.datetime.strptime(p.date, '%y%m%d')
                fmt = ('%d/%m/%Y')
                dateStr = date.strftime(fmt)
                cells[headings.index("Date")].value = dateStr

                labelString = 'Hydrate oil (Batch %d)' % batch
                htask = p.returnTaskByLabel(labelString)
                cells[headings.index("Oil/water vortex time")].value = htask.duration

                labelString = 'Mix with 5%% ABIL (Batch %d)' % batch
                stask = p.returnTaskByLabel(labelString)
                cells[headings.index("final oil/surfactant mix conc %w?")].value = stask.surfactantConcn

                gws.update_cells(cells)

    
    #webbrowser.open('https://docs.google.com/spreadsheets/d/1W_S4NUgCKchcfokpm7cbRywsh-AIfmzk5ksjX05vKII/edit#gid=1149687153', 2, True)

def crossRefPrepToGS(p, gws):

    headings = gws.row_values(1)
    oilIDsFromGS = gws.col_values(headings.index("oil/surfactant batch ID") + 1)
    oilIDsFromGS = oilIDsFromGS[1:]
    sampleNamesFromGS = gws.col_values(headings.index("Sample name") + 1)
    sampleNamesFromGS = sampleNamesFromGS[1:]
    
    print('oil IDs in prep spreadsheet')
    print(p.oilIDs)

    for oilIDFromGS in oilIDsFromGS:
        if (oilIDFromGS in p.oilIDs) & (oilIDFromGS is not None):
            # identify batch corresponding to this ID
            # get final surfactant concn, hydration time directly from p
            # work out (approx.?) rest time based on start print time, oil prep durations, and oil/water mix date and time
            headings = gws.row_values(1)
            row = oilIDsFromGS.index(oilIDFromGS)
            first_col = numberToLetters(headings.index("Protocol version") + 1)
            last_col = numberToLetters(headings.index("Door") + 1) 
            cellrange = '%s%d:%s%d' % (first_col, row+2, last_col, row+2)
            cells = gws.range(cellrange)
            headings = headings[headings.index("Protocol version"):(headings.index("Door") + 1)]
            
            correspondingSampleName = sampleNamesFromGS[row]
            batch = p.oilIDs.index(oilIDFromGS) + 1
            labelString = 'Hydrate oil (Batch %d)' % batch
            htask = p.returnTaskByLabel(labelString)
            cells[headings.index("Oil/water vortex time")].value = htask.duration
            labelString = 'Mix with 5%% ABIL (Batch %d)' % batch
            stask = p.returnTaskByLabel(labelString)
            cells[headings.index("final oil/surfactant mix conc %w?")].value = stask.surfactantConcn
            gws.update_cells(cells)
        
    webbrowser.open('https://docs.google.com/spreadsheets/d/1W_S4NUgCKchcfokpm7cbRywsh-AIfmzk5ksjX05vKII/edit#gid=1149687153', 2, True)

def returnBatchNumber(taskLabel):
    # would be preferable to make this a method of the PrintingPrep class, but that's problematic to call from within class?!
    strs = re.split('\\(|\\)', taskLabel)
    batchstr = strs[1];
    m = re.match('batch (?P<_0>\d+)', batchstr.lower())
    intValue = int(m.group(1))
    return intValue;
            
if __name__ == "__main__":

    try:
        #mode = MODE_NEW
        print(len(sys.argv))
        if len(sys.argv) == 1:
            mode = MODE_NEW
            print('Mode = generate new Excel spreadsheet summary')
        elif len(sys.argv) == 2 :
            mode = MODE_UPDATEWEBFROMCL
            checklistPath = sys.argv[1]
            print('Mode = update google sheet')
        elif len(sys.argv) == 3:
            mode = MODE_APPEND
            checklistPath = sys.argv[1]
            outputFile = sys.argv[2]
            print('Mode = append to existing Excel sheet')
        else:
            errorHandler('Too many arguments passed!')

        #mode = MODE_UPDATEWEBFROMCL
        #mode = MODE_NEW

        xml_export = False;

        if (mode == MODE_UPDATEWEBFROMCL):
        
            gc = authenticate_google_docs()
            gsh = gc.open("Sample register 2016")
            gws = gsh.worksheet("Sample register")
            #gsh = gc.open("Dummy sample register")
            #gws = gsh.worksheet("Sheet1")

            # Replace this with argument input from command line - get from excel using Application.ActiveWorkbook.Path or Application.ActiveWorkbook.FullName 
            #checklistPath = '//base4share/share/SOPs/Completed Checklists/Data/Printing/Printing 1 2015-10-30 0909.xlsm'
            #checklistPath = '//base4share/share/SOPs/Completed Checklists/Data/Printing/Printing 1 2015-11-04 1732.xlsm'
            #checklistPath = '//base4share/share/Doug/P1-151102-A - Copy.xlsm'

            wb = load_workbook(checklistPath)
            ws = wb.active

            dummy = os.path.split(checklistPath)
            print(dummy)
            testString = dummy[0].split('\\Data\\')
            print(testString)

            if (testString[1] == 'Printing'):
                p = Printing(checklistPath)
                p.populatePrintingClass(ws)
                p.populateTasks(ws)
                if (isinstance(p, Printing)):
                    if (p.sampleName != None):
                        if ("ppl" not in p.sampleName.lower()):
                            formatToGS(p, gws)
            else:
                p = PrintingPrep(checklistPath)
                p.populatePrintingPrepClass(ws)
                p.populateTasks(ws)
                # cross reference with Printing sheet here!
                #crossRefPrepToGS(p, gws)
                uploadToHiddenGS(p, gsh)

            

        if (mode == MODE_APPEND):
            print("Mode = APPEND")
            # prompt user to choose type of checklist being summarised, or do so automatically?
            # prompt user for output file, or take from arguments to allow scheduled running?
            # identify output file
            # identify date of last entry
            # identify checklists since last entry
            # loop through these checklists and add a checklist class for each case to a List
            # from output file column titles identify the fields to export - if fields don't exist, add warning text to these entries?
            # from output file, determine list of ID fields that can be used to confirm that new entries don't already exist
            # perform check for exisiting entries and delete those classes from the List
            # loop through classes and fields and parse to output format
            # append to output file - first checking that write is possible and warning user (email?) if not
            # TODO: add option for googlesheets export
            # OPTION: re-export checklist data in XML format?
            print("nonsense")


        if (mode == MODE_NEW):
            print("Mode = NEW")
            # prompt user for output file name/location
            # TODO: add googledocs option?
            initialDir = "//base4share/share/SOPs/Completed Checklists/Data/Printing"
        
            # prompt user for place to look for checklists
            prompt = "Please choose a location in which to look for checklist spreadsheets..."
            #inputPath = chooseFolder(initialDir, prompt)
            #inputPath = "//base4share/share/SOPs/Completed Checklists/Data/Printing"
            inputPath = "C:/Users/d.kelly/Desktop/test/Data/General Printing Prep"

            # generate list of checklist paths
            checklistList = []
            for root, dirs, files in os.walk(inputPath):
                for basename in files:
                    if CHECKLIST_FILE_FORMAT in basename:
                        checklistList.append(os.path.join(root, basename))
            #print(checklistList)
            if (len(checklistList) == 0):
                errorHandler(FILES_LENGTH_ZERO)
        
            # prompt user for checklist type - or get from place to look for checklists?
            # for now, assume that data is in Z:\SOPs\Completed Checklists\Data and figure out which class to use from which folder is selected immediately below
            internalDataList = []
            dummy = os.path.split(checklistList[0])
            testString = dummy[0].split('/Data/')
            print(testString)
            
            for checklist in checklistList:
                print(checklist)
                wb = load_workbook(checklist)
                ws = wb.active
                #try:
                p = []
                if (testString[1] == 'Printing'):
                    print("Use Printing class")
                    p = Printing(checklist)
                    p.populatePrintingClass(ws)
                    
                elif (testString[1] == 'General Printing Prep'):
                    print("Use PrintingPrep class");
                    p = PrintingPrep(checklist)
                    p.populatePrintingPrepClass(ws)
                    
                else:
                    errorHandler(NOT_YET_SUPPORTED)
                
                p.populateTasks(ws)
                if (isinstance(p, Printing)):
                    if (p.sampleName != None):
                        if ("ppl" not in p.sampleName.lower()):
                            internalDataList.append(p)
                else:
                    internalDataList.append(p)

            # TODO: better to establish correct number of tasks by adding tasks to a dictionary so that
            # there is a list of unique tasks. 
            no_tasks = []
            for internalData in internalDataList:
                no_tasks.append(len(internalData.tasks))


            # prompt user for fields to include in summary      
            m = outputSelectionDialog(internalDataList[no_tasks.index(max(no_tasks))])
            m.root.mainloop()
            m.root.destroy()
            print('run past dialog, data to save:')
            for t in m.data.tasks:
                if (t.output):
                    print(t.taskLabel)

            # loop through classes and fields and parse to output format
            # set up file to output to
            if isinstance(m.data, PrintingPrep):
                descString = "printing prep"
            else:
                descString = "printing"

            initialFile = "%s %s data summary.xlsx" % (time.strftime('%Y-%m-%d'), descString)
            outputFile = chooseOutputFile(initialDir, initialFile)
            #outputFile = "C:/Users/d.kelly/Desktop/Python/DKBase4PythonScripts/checklistDataMining/sampleData/test/output.xlsx"
            wb = Workbook()
            ws = wb.active

            exclude_vars = ['taskLabel', 'taskCategory', 'taskNumber', 'notes', 'output', 'outerInstance', 'batch']
        
            # set up invariant section
            r = 1
            if (isinstance(internalData, Printing)):
                ws.cell(row = r, column = 1).value = 'Source checklist'
                ws.cell(row = r, column = 2).value = 'Print date'
                ws.cell(row = r, column = 3).value  = 'Sample name'
                ws.cell(row = r, column = 4).value  = 'Print rig'
                ws.cell(row = r, column = 5).value  = 'Experimenter'
                ws.cell(row = r, column = 6).value  = 'QC'
                ws.cell(row = r, column = 7).value  = 'SOP version'

                for internalData in internalDataList:
                    r = r + 1
                    ws.cell(row = r, column = 1).value = internalData.path
                    ws.cell(row = r, column = 2).value = internalData.printDate
                    ws.cell(row = r, column = 3).value  = internalData.sampleName
                    ws.cell(row = r, column = 4).value  = internalData.printRig
                    ws.cell(row = r, column = 5).value  = internalData.experimenter
                    ws.cell(row = r, column = 6).value  = internalData.qCPerson
                    ws.cell(row = r, column = 7).value  = internalData.sOPVersion

                col = 8
            else:
                ws.cell(row = r, column = 1).value = 'Source checklist'
                ws.cell(row = r, column = 2).value  = 'Date'
                ws.cell(row = r, column = 3).value  = 'QC'
                ws.cell(row = r, column = 4).value  = 'SOP version'

                for internalData in internalDataList:
                    r = r + 1
                    ws.cell(row = r, column = 1).value = internalData.path
                    ws.cell(row = r, column = 2).value = internalData.date
                    ws.cell(row = r, column = 3).value  = internalData.qCPerson
                    ws.cell(row = r, column = 4).value  = internalData.sOPVersion

                col = 5

            for mt in m.data.tasks:
                if mt.output:
                    v = vars(mt)
                    for key in v:
                        if key not in exclude_vars:
                            r = 1
                            col_filled = False
                            ws.cell(row = r, column = col).value = "%s: %s" % (mt.taskLabel, key)
                            print(mt.taskLabel)
                            for internalData in internalDataList:
                                r = r + 1
                                for task in internalData.tasks:
                                    if (task.taskLabel == mt.taskLabel) and (task.taskCategory == mt.taskCategory):
                                        vv = vars(task)
                                        if (vv[key] != None) and (vv[key] != ""):
                                            col_filled = True
                                        ws.cell(row = r, column = col).value = vv[key]
                            if col_filled:
                                col = col + 1

            # write to output file (googledoc?)
            wb.save(outputFile)
    except:
        errorHandler(traceback.format_exc())


